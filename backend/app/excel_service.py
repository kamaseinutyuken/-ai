import os
import pandas as pd
import tempfile
from typing import Dict, Any, List, Tuple
from openpyxl import load_workbook

SAFETY_PLAN_RANGES = [
    (34, 43),   # Row ranges
    (54, 73),
    (81, 100),
    (108, 127)
]

SAFETY_PLAN_COLUMNS = ["D", "AS", "BK"]

RISK_ASSESSMENT_COLUMNS = ["C", "O", "Y", "AI", "AS", "CC", "AV", "CF", "BE", "CL"]

RISK_ASSESSMENT_ROWS = [
    34, 37, 40, 48, 51, 54, 57, 60, 63, 66, 69, 72, 75, 78, 81, 84, 87, 96, 99, 102,
    105, 108, 111, 114, 117, 120, 123, 126, 129, 132, 135, 144, 147, 150, 153, 156,
    159, 162, 165, 168, 171, 174, 177, 180, 183, 192, 195, 198, 201, 204, 207, 210,
    213, 216, 219, 222, 225, 228, 231, 240, 243, 246, 249, 252, 255, 258, 261, 264,
    267, 270, 273, 276, 279, 288, 291, 294, 297, 300, 303, 306, 309, 312, 315, 318,
    321, 324, 327, 336, 339, 342, 345, 348, 351, 354, 357, 360, 363, 366, 369, 372,
    375, 384, 387, 390, 393, 396, 399, 402, 405, 408, 411, 414, 417, 420, 423, 432,
    435, 438, 441, 444, 447, 450, 453, 456, 459, 462, 465, 468, 471, 480, 483, 486,
    489, 492, 495, 498, 501, 504, 507, 510, 513, 516, 519, 528, 531, 534, 537, 540,
    543, 546, 549, 552, 555, 558, 561, 564, 567
]

def convert_excel_col_to_index(col_str: str) -> int:
    """Convert Excel column string to index (0-based)"""
    result = 0
    for char in col_str:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1

def save_uploaded_file(file_content: bytes, filename: str) -> str:
    """Save uploaded file to a temporary location and return the path"""
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)
    
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    return file_path

def extract_knowledge_from_excel(file_path: str) -> Dict[str, Any]:
    """
    Extract knowledge from Excel file based on specified cell ranges
    Returns a dictionary with structured data from the Excel file
    """
    try:
        wb = load_workbook(file_path)
        
        required_sheets = ["安全施工計画書", "リスクアセスメント"]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Required sheet '{sheet_name}' not found in Excel file")
        
        safety_plan_data = {}
        safety_plan_sheet = wb["安全施工計画書"]
        
        fixed_cells = {
            "I9": "工事名",
            "I11": "施工場所",
            "I13": "工期",
            "Q15": "作業者数"
        }
        
        fixed_cell_values = {}
        for cell, description in fixed_cells.items():
            value = safety_plan_sheet[cell].value
            fixed_cell_values[description] = value
        
        safety_plan_data["fixed_cells"] = fixed_cell_values
        
        range_data = []
        for col in SAFETY_PLAN_COLUMNS:
            col_data = {}
            for start_row, end_row in SAFETY_PLAN_RANGES:
                row_data = {}
                for row in range(start_row, end_row + 1):
                    cell_ref = f"{col}{row}"
                    value = safety_plan_sheet[cell_ref].value
                    if value is not None and value != "":
                        row_data[row] = value
                if row_data:
                    col_data[f"{start_row}-{end_row}"] = row_data
            range_data.append({"column": col, "data": col_data})
        
        safety_plan_data["range_data"] = range_data
        
        risk_assessment_data = {}
        risk_sheet = wb["リスクアセスメント"]
        
        column_data = {}
        for col in RISK_ASSESSMENT_COLUMNS:
            col_data = {}
            for row in RISK_ASSESSMENT_ROWS:
                cell_ref = f"{col}{row}"
                value = risk_sheet[cell_ref].value
                if value is not None and value != "":
                    col_data[row] = value
            if col_data:
                column_data[col] = col_data
        
        risk_assessment_data["column_data"] = column_data
        
        knowledge_base = {
            "safety_plan": safety_plan_data,
            "risk_assessment": risk_assessment_data
        }
        
        return knowledge_base
    
    except Exception as e:
        raise Exception(f"Error extracting knowledge from Excel: {str(e)}")

def analyze_excel_with_pandas(file_path: str) -> Dict[str, Any]:
    """
    Analyze Excel file using pandas for additional insights
    Returns statistical information about the Excel data
    """
    try:
        xls = pd.ExcelFile(file_path)
        
        analysis_results = {}
        
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            sheet_stats = {
                "rows": df.shape[0],
                "columns": df.shape[1],
                "non_empty_cells": df.count().sum(),
                "empty_cells": df.isna().sum().sum(),
            }
            
            column_stats = {}
            for col in df.columns:
                col_stats = {
                    "non_empty": df[col].count(),
                    "empty": df[col].isna().sum(),
                    "unique_values": df[col].nunique(),
                }
                column_stats[str(col)] = col_stats
            
            sheet_stats["column_stats"] = column_stats
            analysis_results[sheet_name] = sheet_stats
        
        return analysis_results
    
    except Exception as e:
        raise Exception(f"Error analyzing Excel with pandas: {str(e)}")

def combine_excel_analysis(file_path: str) -> Dict[str, Any]:
    """
    Combine both knowledge extraction and pandas analysis
    """
    knowledge_base = extract_knowledge_from_excel(file_path)
    pandas_analysis = analyze_excel_with_pandas(file_path)
    
    return {
        "knowledge_base": knowledge_base,
        "statistical_analysis": pandas_analysis
    }
