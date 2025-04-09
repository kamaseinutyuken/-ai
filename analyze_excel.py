import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

excel_file_path = os.path.expanduser("~/attachments/c49c933d-c879-4bc7-9820-863924d20bc0/.xlsx")

try:
    wb = load_workbook(excel_file_path)
    
    print("Sheet names in the Excel file:")
    for sheet_name in wb.sheetnames:
        print(f"- {sheet_name}")
    
    required_sheets = ["安全施工計画書", "リスクアセスメント"]
    for sheet_name in required_sheets:
        if sheet_name in wb.sheetnames:
            print(f"\nFound required sheet: {sheet_name}")
        else:
            print(f"\nWARNING: Required sheet '{sheet_name}' not found!")
    
    print("\nChecking specific cells:")
    
    if "安全施工計画書" in wb.sheetnames:
        sheet = wb["安全施工計画書"]
        
        fixed_cells = {
            "I9": "工事名",
            "I11": "施工場所",
            "I13": "工期",
            "Q15": "作業者数"
        }
        
        for cell, description in fixed_cells.items():
            value = sheet[cell].value
            print(f"{description} ({cell}): {value}")
        
        sample_cells = ["D34", "AS34", "BK34", "D54", "AS54", "BK54"]
        print("\nSample cells from 安全施工計画書:")
        for cell in sample_cells:
            value = sheet[cell].value
            print(f"{cell}: {value}")
    
    if "リスクアセスメント" in wb.sheetnames:
        sheet = wb["リスクアセスメント"]
        
        sample_cells = ["C34", "O34", "Y34", "AI34", "AS34", "CC34"]
        print("\nSample cells from リスクアセスメント:")
        for cell in sample_cells:
            value = sheet[cell].value
            print(f"{cell}: {value}")
    
    print("\nSheet dimensions:")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"{sheet_name}: {sheet.dimensions}")

except Exception as e:
    print(f"Error analyzing Excel file: {e}")
